#!/usr/bin/python
#coding:utf-8
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook

import sys,os
import re
import json
import logging
reload(sys)#need reload sys
sys.setdefaultencoding('utf-8')

path_svn_ui= 'svn://10.128.161.92/datacard/QCT_MDM9x30/branches/WEBUI_SRC/SmartSwReleaseUI/'
path_svn_ui_string= 'svn://10.128.161.92/datacard/QCT_MDM9x30/branches/WEBUI_SRC/'
path_svn_spm= 'svn://10.128.161.92/datacard/SPM/03" "Customized" "Requirements'
dir_webrc= 'webrc'#file name
generic_name= 'Generic'#default generic name
project_name= 'MW40'#default project name
custom_name= 'www'#default custom name
sys_language= ['en','fr']


def run_sys_cmd(sys_cmd):
    print sys_cmd
    os.system(sys_cmd)

def add_key(obj,n,v):
    if('Null' == v):
        return
    else:
        obj[n]= v+'\n'


def export_ui_code():
    run_sys_cmd('svn export --force %s' %(path_svn_ui+'/'+generic_name+'/www/'))
    run_sys_cmd('svn export --force %s' %(path_svn_ui+'/'+generic_name+'/'+project_name+'_rc'))
    run_sys_cmd('svn export --force %s' %(path_svn_ui+'/'+generic_name+'/'+custom_name))

def copy_ui_code():
    if os.path.exists(dir_webrc):
        run_sys_cmd('rm -fr %s' %(dir_webrc))
    run_sys_cmd('mkdir %s' %(dir_webrc))
    run_sys_cmd('cp -rf %s/* %s/' %('www', dir_webrc))
    run_sys_cmd('''cp -rf %s/* %s/''' %(project_name+'/_rc', dir_webrc))
    run_sys_cmd('''cp -rf %s/* %s/''' %(custom_name+'/www', dir_webrc))
    run_sys_cmd('''cp -rf %s/* %s/''' %(custom_name+'/'+project_name+'_rc', dir_webrc))

def delete_ui_file():
    run_sys_cmd('''rm -rf %s''' %('www'))
    run_sys_cmd('''rm -rf %s''' %(custom_name))
    run_sys_cmd('''rm -rf %s''' %(project_name+'_rc'))


def export_requireSheet():
    run_sys_cmd('''rm -rf %s''' %('Requirement-all.xlsx'))
    run_sys_cmd('''rm -rf %s''' %('Requirement-all.xlsm'))
    run_sys_cmd('svn export --force %s' %(path_svn_spm+'/'+requireSheet_path))

color_path= ''
def export_colorSheet():
    print color_path
    new_color_path= searchRepalce(color_path).replace('\\','/').replace(' ','" "')
    run_sys_cmd('''rm -rf %s''' %('colour_all.xlsm'))
    run_sys_cmd('svn export --force %s' %(path_svn_spm+'/'+new_color_path))

def export_um():
    print um_path
    new_um_path= searchRepalce(um_path).replace('\\','/').replace(' ','" "')
    #print um_path.split('\\')[len(um_path.split('\\')) -2]
    run_sys_cmd('svn export --force %s' %(path_svn_spm+'/UM/'+new_um_path))
    run_sys_cmd('cp -rf %s/* %s/' %(um_path.split('\\')[len(um_path.split('\\')) -2], dir_webrc+'/help'))
    run_sys_cmd('''rm -rf %s''' %(um_path.split('\\')[len(um_path.split('\\')) -2]))

def export_ui_stringSheet():
    run_sys_cmd('''rm -rf %s''' %('all_string.xlsx'))
    run_sys_cmd('svn export --force %s' %(path_svn_ui_string+'/Strings/all_string.xlsx'))


map_string_obj= {}
map_ids= 'ids'
map_strID= 'strID'
def parse_stringSheet():
    global map_string_obj
    print 'Info: parse_string_sheet'
    if not os.path.exists(dir_webrc+'/all_map.xlsx'):
        print 'Error: all_map.xlsx'
        return False
    if not os.path.exists('all_string.xlsx'):
        print 'Error: all_string.xlsx'
        return False
    wb= load_workbook(filename=os.getcwd()+'/'+dir_webrc+'/all_map.xlsx',read_only=True)
    ws= wb['maps']
    for row in ws.rows:
        n= 0
        for cell in row:
            n+= 1
            if None != cell.value:
                if searchRepalce(cell.value) == map_ids:
                    map_ids_index= n
                if searchRepalce(cell.value) == map_strID:
                    map_strID_index= n
        break

    for row in ws.rows:
        m= 0
        cache_obj= {}
        for cell in row:
            m+= 1
            if None != cell.value:
                if map_ids_index == m:
                    add_key(cache_obj, map_ids, searchRepalce(cell.value))
                if map_strID_index == m:
                    add_key(cache_obj, map_strID, searchRepalce(cell.value))
        if cache_obj.has_key(map_ids) and cache_obj.has_key(map_strID):
            add_key(map_string_obj, searchRepalce(cache_obj[map_ids]), searchRepalce(cache_obj[map_strID]))

    if os.path.exists(dir_webrc+'/custom_map.xlsx'):
        c_wb= load_workbook(filename=os.getcwd()+'/'+dir_webrc+'/custom_map.xlsx',read_only=True)
        c_ws= c_wb['maps']
        for row in c_ws.rows:
            c= 0
            for cell in row:
                c+= 1
                if None != cell.value:
                    if searchRepalce(cell.value) == map_ids:
                        c_map_ids_index= c
                    if searchRepalce(cell.value) == map_strID:
                        c_map_strID_index= c
            break

        for row in c_ws.rows:
            d= 0
            c_cache_obj= {}
            for cell in row:
                d+= 1
                if None != cell.value:
                    if c_map_ids_index == d:
                        add_key(c_cache_obj, map_ids, searchRepalce(cell.value))
                    if c_map_strID_index == d:
                        add_key(c_cache_obj, map_strID, searchRepalce(cell.value))
            if c_cache_obj.has_key(map_ids) and c_cache_obj.has_key(map_strID):
                add_key(map_string_obj, searchRepalce(c_cache_obj[map_ids]), searchRepalce(c_cache_obj[map_strID]))
        c_wb.close()
    wb.close()
    parse_all_string()

StrID= 'StrID'
def parse_all_string():
    wb= load_workbook(filename=os.getcwd()+'/all_string.xlsx',read_only=True)
    ws= wb['Strings']
    n= 0
    for lang in sys_language:
        n+= 1
        write_string(wb, ws, lang)
    if len(sys_language) == n:
        wb.close()

all_string_obj= {}
def write_string(wb, ws, language):
    for row in ws.rows:
        n= 0
        for cell in row:
            n+= 1
            if None != cell.value:
                if searchRepalce(cell.value) == StrID:
                    StrID_index= n
                if searchRepalce(cell.value) == language:
                    language_index= n
        break

    for row in ws.rows:
        m= 0
        cache_obj= {}
        for cell in row:
            m+= 1
            if None != cell.value:
                if searchRepalce(cell.value) == StrID:
                    break
                if StrID_index == m:
                    add_key(cache_obj, StrID, searchRepalce(cell.value))
                if language_index == m:
                    add_key(cache_obj, language, searchRepalce(cell.value))
        if cache_obj.has_key(StrID) and cache_obj.has_key(language):
            add_key(all_string_obj,searchRepalce(cache_obj[StrID]), searchRepalce(cache_obj[language]))

    w= open('res_'+language+'.res','w')
    w.write('var resourceInfo={\n')
    i= 0
    dot= ','
    l= len(map_string_obj)
    for k,v in map_string_obj.iteritems():
        i+= 1
        if l == i:
            dot= ''
        if all_string_obj.has_key(searchRepalce(v)):
            w.write('''%s:'%s'%s\n''' %(searchRepalce(k), filterString(searchRepalce(all_string_obj[searchRepalce(v)])), dot))
    #print result_obj
    w.write('}')
    w.close()
    print 'Info: complete string '+language
    run_sys_cmd('cp -rf %s %s/' %('res_'+language+'.res',dir_webrc+'/resource'))
    run_sys_cmd('rm -rf %s' %('res_'+language+'.res'))


def filterString(value):
    return str(value).replace('\\','\\\\').replace('\'','\\\'')


color_obj= {}
colour_ID= "colourID"
default_value= "defaultValue"
colour_vale= "colourVale"
def parse_colorSheet():
    global color_obj
    if not os.path.exists('colour_all.xlsm'):
        print 'Error: colour_all.xlsm'
        return False
    wb= load_workbook(filename=os.getcwd()+'/colour_all.xlsm',read_only=True)
    ws= wb['colour_v1.0']
    for row in ws.rows:
        n= 0
        for cell in row:
            n+= 1
            if searchRepalce(cell.value) == colour_ID:
                colour_ID_index= n
            if searchRepalce(cell.value) == default_value:
                default_value_index= n
            if searchRepalce(cell.value) == colour_vale:
                colour_vale_index= n
        break
    #print colour_ID_index, default_value_index, colour_vale_index
    for row in ws.rows:
        n= 0
        cache_obj= {}
        for cell in row:
            if colour_ID == cell.value:
                break
            n+= 1
            if None != cell.value:
                if colour_ID_index == n:
                    add_key(cache_obj, colour_ID, searchRepalce(cell.value))
                if default_value_index == n and searchRepalce(cell.value) != None:
                    add_key(cache_obj, default_value, searchRepalce(cell.value))
                if colour_vale_index == n and searchRepalce(cell.value) != None:
                    add_key(cache_obj, colour_vale, searchRepalce(cell.value))
        if cache_obj.has_key(colour_ID) and cache_obj.has_key(colour_vale):
            add_key(color_obj, searchRepalce(cache_obj[colour_ID]), searchRepalce(cache_obj[colour_vale]))
        elif cache_obj.has_key(colour_ID) and cache_obj.has_key(default_value):
            add_key(color_obj, searchRepalce(cache_obj[colour_ID]), searchRepalce(cache_obj[default_value]))
    #print color_obj
    wb.close()

def write_color():
    w= open('color.css','w')
    o= open(dir_webrc+'/css/color.css','r')
    lines= o.readlines()
    for line in lines:
        matchObj= re.match(r'(.*)(\$.*);', line, re.M|re.I)
        if matchObj:
            if color_obj.has_key(matchObj.group(2)):
                line= line.replace(matchObj.group(2), searchRepalce(color_obj[matchObj.group(2)]))
            else:
                print matchObj.group(2)
        w.write(line)
    o.close()
    w.close()
    run_sys_cmd('cp -rf %s/* %s' %('color.css',dir_webrc+'/css/'))
    run_sys_cmd('rm -rf %s' %('color.css'))
    run_sys_cmd('rm -rf %s' %('colour_all.xlsm'))
    print 'Info: write successful'



table_name= 'Table name'
key_name= 'Key name'
js_key= 'Js key'
def parse_requireSheet():
    global generic_name, project_name, color_path, um_path, sys_language
    if os.path.exists('Requirement-all.xlsx'):
        format_ssuffix= 'Requirement-all.xlsx'
    elif os.path.exists('Requirement-all.xlsm'):
        format_ssuffix= 'Requirement-all.xlsm'
    else:
        format_ssuffix= ''
        print 'Error: Requirement-all.xlsm or Requirement-all.xlsx'
        return False

    if format_ssuffix != '':
        wb= load_workbook(filename=os.getcwd()+'/'+format_ssuffix,read_only=True)
        ws= wb['SW-V1.0']
        first_row= {}
        for row in ws.rows:
            n= 0
            for cell in row:
                n+= 1
                if None != cell.value:
                    if custom_name == searchRepalce(cell.value):
                        custom_name_index= n
                    if table_name == searchRepalce(cell.value):
                        table_name_index= n
                    if key_name == searchRepalce(cell.value):
                        key_name_index= n
                    if js_key == searchRepalce(cell.value):
                        js_key_index= n
            break
        print custom_name_index, table_name_index, key_name_index
        row_index= ''
        m= 0
        for row in ws.rows:
            n= 0
            m+= 1
            row_info= {}
            for cell in row:
                n+= 1
                if None != cell.value: 
                    if table_name_index == n and searchRepalce(cell.value).find("_ui_") == -1 and searchRepalce(cell.value).find("_um_") == -1:
                        break
                    if searchRepalce(cell.value) == 'ProjectName':
                        row_index= 'pn'
                    if searchRepalce(cell.value) == 'WebUISytle':
                        row_index= 'wuis'
                    if searchRepalce(cell.value) == 'ColourXlsxPath':
                        row_index= 'cxp'
                    if searchRepalce(cell.value) == '_um_':
                        row_index= 'um'
                    if searchRepalce(cell.value) == 'sys_language':
                        row_index= 'lang'
                    add_key(row_info, n, searchRepalce(cell.value))
            if row_info.has_key(custom_name_index) and row_index == 'pn':
                project_name= searchRepalce(row_info[custom_name_index])
                row_index= ''
            elif row_info.has_key(custom_name_index) and row_index == 'wuis':
                generic_name= searchRepalce(row_info[custom_name_index])
                row_index= ''
            elif row_info.has_key(custom_name_index) and row_index == 'cxp':
                color_path= searchRepalce(row_info[custom_name_index])
                row_index= ''
            elif row_info.has_key(custom_name_index) and row_index == 'um':
                um_path= searchRepalce(row_info[custom_name_index])
                row_index= ''
            elif row_info.has_key(custom_name_index) and row_index == 'lang':
                sys_language= searchRepalce(row_info[custom_name_index]).split(',')
                row_index= ''
        print 'Info: generic='+generic_name
        print 'Info: project='+project_name
        print 'Info: custom='+custom_name
        print 'Info: language='
        print sys_language
        print 'Info: color='+color_path
        print 'Info: um='+um_path
        parse_configAuto(wb, ws, custom_name_index, table_name_index, js_key_index)

js_key_obj= {}
jsKey= 'key'
jsKey_value= 'value'
table_name_index= '_ui_'
def parse_configAuto(wb, ws, custom_name_index, table_name_index, js_key_index):
    global js_key_obj
    for row in ws.rows:
        n= 0
        jsKeyValue_obj= {}
        for cell in row:
            n+= 1
            if None != cell.value:
                if table_name_index == n and searchRepalce(cell.value).find("_ui_") == -1:
                    break
                if js_key_index == n and searchRepalce(cell.value)  != None:
                    add_key(jsKeyValue_obj, jsKey, searchRepalce(cell.value))
                if custom_name_index == n and searchRepalce(cell.value) != None:
                    add_key(jsKeyValue_obj, jsKey_value, searchRepalce(cell.value))
        if jsKeyValue_obj.has_key(jsKey) and jsKeyValue_obj.has_key(jsKey_value):
            add_key(js_key_obj,jsKeyValue_obj[jsKey], jsKeyValue_obj[jsKey_value])
    #print js_key_obj
    wb.close()
    write_configAuto(js_key_obj)

def write_configAuto(js_key_obj):
    w= open('configAuto.js','w')
    w.write('var configAuto = {\n')
    dot= ','
    n= 0
    l= len(js_key_obj)
    for  k,v in js_key_obj.iteritems():
        n+= 1
        if l == n:
            dot= ''
        if 'sys_language' == searchRepalce(k):
            w.write('''\t%s:%s%s\n''' %(searchRepalce(k),searchRepalce(v).split(','),dot))
        elif None != searchRepalce(v) and searchRepalce(v).isdigit():
            w.write('''\t%s:%s%s\n''' %(searchRepalce(k),searchRepalce(v),dot))
        else:
            w.write('''\t%s:"%s"%s\n''' %(searchRepalce(k),searchRepalce(v),dot))
    w.write('}')
    w.close()
    run_sys_cmd('cp -rf %s/ %s' %('configAuto.js',dir_webrc+'/js/'))
    run_sys_cmd('rm -rf %s' %('configAuto.js'))


def searchRepalce(strings):
    return re.sub(r'\n', '', str(strings)).strip()

def loggingInfo(info):
    logging.info(info)


def main(input_list):
    global requireSheet_path, custom_name
    #print input_list
    python_copy_file= input_list.pop(0)#del python_copy.py
    print input_list
    custom_name= input_list[0].replace(' ','" "')
    requireSheet_path= input_list[1]
    export_requireSheet()
    parse_requireSheet()
    export_colorSheet()
    export_um()
    export_ui_code()
    copy_ui_code()
    delete_ui_file()
    parse_colorSheet()
    write_color()
    export_ui_stringSheet()
    parse_stringSheet()


if __name__ == '__main__':
    print 'Info: --------------Start copy code-------------------'
    if (len(sys.argv) != 3):
        print '-------Error: eg:python python_copy.py custom_name requireSheetPath--------------------'
    else:
        main(sys.argv)